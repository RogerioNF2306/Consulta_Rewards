import asyncio
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import quote

import openpyxl
from dotenv import load_dotenv
from playwright.async_api import async_playwright

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

try:
    from modulo_links.config import load_config
    from modulo_links.excel_manager import carregar_planilha, salvar_planilha, fechar_excel_forcado
    from modulo_links.navegador_links import criar_contexto, buscar_no_xbox, interacao_selecao
    from modulo_links.utils import USER_DATA_DIR
except ImportError:
    from config.config import load_config
    from excel_manager import carregar_planilha, salvar_planilha, fechar_excel_forcado
    from navegador_links import criar_contexto, buscar_no_xbox, interacao_selecao
    from config.utils import USER_DATA_DIR

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(ROOT_DIR, '.env'))

EXCEL_PATH = os.getenv("EXCEL_PATH", "jogos.xlsx")
EXCEL_SHEET = os.getenv("EXCEL_SHEET", "Sheet1")

if not os.path.isabs(EXCEL_PATH):
    EXCEL_PATH = os.path.join(ROOT_DIR, EXCEL_PATH)


def limpar_nome_jogo(nome):
    if not nome:
        return ""
    nome = str(nome)
    nome = re.sub(r'[™®©]', '', nome)
    nome = re.sub(r'\(.*?\)', '', nome)
    return " ".join(nome.split())


def link_valido(link, config):
    if not link:
        return False
    return bool(re.search(config["valid_link_pattern"], str(link)))


def fechar_excel_forcado():
    try:
        print("\n⏳ Solicitando encerramento forçado do Excel...")
        os.system("taskkill /f /im excel.exe >nul 2>&1")
        time.sleep(1.5)
        print("✅ Processo do Excel encerrado com sucesso.")
    except Exception as e:
        print(f"⚠️ Não foi possível fechar o Excel automaticamente: {e}")


async def salvar_alteracao(linha, coluna, valor):
    wb_write = carregar_planilha(EXCEL_PATH)
    if wb_write is None:
        return False
    try:
        ws_write = wb_write[EXCEL_SHEET]
        ws_write.cell(row=linha, column=coluna).value = valor
        return salvar_planilha(wb_write, EXCEL_PATH)
    except Exception as e:
        print(f"❌ Erro ao atualizar célula em segundo plano: {e}")
        try:
            wb_write.close()
        except Exception:
            pass
        return False


async def executar_modo(modo):
    wb_leitura = carregar_planilha(EXCEL_PATH)
    if wb_leitura is None:
        return

    try:
        ws = wb_leitura[EXCEL_SHEET]
        jogos_planilha = []
        for r in range(2, ws.max_row + 1):
            n = ws.cell(row=r, column=1).value
            l = ws.cell(row=r, column=2).value
            if n:
                jogos_planilha.append({"row": r, "nome": n, "link": l})
    except KeyError:
        print(f"❌ Aba '{EXCEL_SHEET}' não encontrada no arquivo Excel.")
        return
    finally:
        wb_leitura.close()

    async with async_playwright() as p:
        context = await criar_contexto(p, USER_DATA_DIR)
        page = context.pages[0]
        config = load_config()

        if modo == 's':
            filtro = ""
            while True:
                lista_exibicao = [j for j in jogos_planilha if filtro.lower() in str(j['nome']).lower()]
                if not lista_exibicao:
                    print("\n⚠️ Nenhum jogo encontrado com o filtro atual.")
                else:
                    print("\n" + f"📑 SELEÇÃO DE JOGOS {'[FILTRO: ' + filtro + ']' if filtro else ''} " + "─"*30)
                    for i, jogo in enumerate(lista_exibicao):
                        status = "✅" if link_valido(jogo['link'], config) else "❌"
                        print(f"{i+1}. {status} {jogo['nome']}")

                print("\n" + "─"*50)
                print("[F] Filtrar | [T] Limpar Filtro | [0] Voltar")
                cmd = input("👉 Escolha: ").strip().lower()
                if cmd == '0':
                    break
                if cmd == 't':
                    filtro = ""
                    continue
                if cmd == 'f':
                    filtro = input("🔎 Nome do jogo: ").strip()
                    continue
                if cmd.isdigit() and 1 <= int(cmd) <= len(lista_exibicao):
                    selecionado = lista_exibicao[int(cmd)-1]
                    res = await interacao_selecao(page, selecionado['nome'], selecionado['link'], config)
                    if res == "SAIR":
                        break
                    if res is not None:
                        if await salvar_alteracao(selecionado['row'], 2, res):
                            selecionado['link'] = res
                            print("💾 Salvo com sucesso!")
                    continue
                print("⚠️ Opção inválida. Tente novamente.")

        else:
            busca_filtro = input("\n🔎 Filtrar por nome (Enter para todos): ").strip().lower()
            jogos_proc = [j for j in jogos_planilha if busca_filtro in str(j['nome']).lower()]
            if not jogos_proc:
                print("⚠️ Nenhum jogo encontrado para processar.")
            for i, jogo in enumerate(jogos_proc):
                row, nome_plan, link = jogo['row'], jogo['nome'], jogo['link']
                if modo == 'a' and link_valido(link, config):
                    continue

                print(f"\n[{i+1}/{len(jogos_proc)}] Processando: {nome_plan}")
                if link_valido(link, config):
                    try:
                        await page.goto(link, wait_until="networkidle", timeout=30000)
                        nome_site = await page.inner_text(config['product_title_h1'])
                        if limpar_nome_jogo(nome_plan).lower() == limpar_nome_jogo(nome_site).lower():
                            print(f"✅ Nome condizente: {nome_site}")
                        else:
                            print(f"⚠️ NOME DIVERGENTE! Planilha: {nome_plan} | Site: {nome_site}")
                    except Exception:
                        link = None

                if not link_valido(link, config):
                    res = await interacao_selecao(page, nome_plan, link, config)
                    if res == "SAIR":
                        break
                    if res is not None:
                        if await salvar_alteracao(row, 2, res):
                            print("💾 Link atualizado com sucesso!")

        await context.close()


if __name__ == '__main__':
    modo = ''
    while modo not in ['a', 'v', 's', '0']:
        print("\n" + "═"*45 + "\n      🎮 GERENCIADOR XBOX (MODULAR)\n" + "═"*45)
        print("[A] ATUALIZAR | [V] VERIFICAR | [S] SELECIONAR | [0] VOLTAR")
        modo = input("Selecione: ").strip().lower()
    if modo != '0':
        asyncio.run(executar_modo(modo))
